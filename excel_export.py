# ╔══════════════════════════════════════════════════════════════════════╗
# ║  excel_export.py  —  Excel workbook builders                       ║
# ╚══════════════════════════════════════════════════════════════════════╝

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any


THIN = Side(style="thin", color="CCCCCC")


def _bdr():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _hdr(cell, text, bg, fg="FFFFFF", sz=10):
    cell.value = text
    cell.font  = Font(bold=True, color=fg, name="Calibri", size=sz)
    cell.fill  = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _bdr()


def _val(cell, v, fmt, bg=None):
    cell.value  = v
    cell.number_format = fmt
    cell.font   = Font(name="Calibri", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _bdr()
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)


def _lbl(cell, text, bg="F2F2F2"):
    cell.value = text
    cell.font  = Font(bold=True, name="Calibri", size=10)
    cell.fill  = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = _bdr()


def _tot(cell, v, fmt, bg):
    cell.value  = v
    cell.number_format = fmt
    cell.font   = Font(bold=True, name="Calibri", size=10, color="FFFFFF")
    cell.fill   = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _bdr()


# ──────────────────────────────────────────────────────────────────────
#  SECTION 1  —  Discount Analysis Excel
# ──────────────────────────────────────────────────────────────────────

def build_s1_excel(
    results_df: pd.DataFrame,
    months_ordered: List[str],
    insights: Dict[Any, dict],
    overall_insights: Dict[str, dict],
) -> io.BytesIO:
    wb = Workbook()

    # ── Summary Matrix sheet ────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary Matrix"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 20

    for i, month in enumerate(months_ordered):
        dc = get_column_letter(2 + i * 2)
        nc = get_column_letter(3 + i * 2)
        ws.column_dimensions[dc].width = 22
        ws.column_dimensions[nc].width = 22
        ws.merge_cells(f"{dc}1:{nc}1")
        _hdr(ws[f"{dc}1"], month.upper(), "1F3864", sz=11)
        _hdr(ws[f"{dc}2"], "Discounted",     "2563EB")
        _hdr(ws[f"{nc}2"], "Non-Discounted", "059669")
        _hdr(ws[f"{dc}3"], "Value", "4472C4", sz=9)
        _hdr(ws[f"{nc}3"], "Value", "4472C4", sz=9)

    ws.merge_cells("A1:A3")
    _hdr(ws["A1"], "METRIC", "0F172A", sz=11)
    ws.freeze_panes = "A4"

    metrics = [
        ("Total Spend (INR)",   "Spend",       "#,##0.00"),
        ("Total Revenue (INR)", "Revenue",     "#,##0.00"),
        ("Spend %",             "Spend_Pct",   "0.0%"),
        ("Revenue %",           "Revenue_Pct", "0.0%"),
        ("ROI",                 "ROI",         '0.00"x"'),
    ]
    for r, (label, key, fmt) in enumerate(metrics, start=4):
        ws.row_dimensions[r].height = 20
        _lbl(ws[f"A{r}"], label)
        for i, month in enumerate(months_ordered):
            dc = get_column_letter(2 + i * 2)
            nc = get_column_letter(3 + i * 2)
            dr = results_df[(results_df["Month"] == month) & (results_df["Category"] == "Discounted")].iloc[0]
            nr = results_df[(results_df["Month"] == month) & (results_df["Category"] == "Non-Discounted")].iloc[0]
            _val(ws[f"{dc}{r}"], dr[key], fmt, "D6E4F0")
            _val(ws[f"{nc}{r}"], nr[key], fmt, "E8F5E9")

    # ── Insight sheets ──────────────────────────────────────────────
    def _write_insights_sheet(wb, sname, data_dict, month_label=None):
        pi = wb.create_sheet(sname)
        pi.sheet_view.showGridLines = False
        SC = {
            ("Discounted",     "hslr"): ("C0392B", "FDEDEC", "E74C3C"),
            ("Discounted",     "lshr"): ("1A5276", "D6EAF8", "2874A6"),
            ("Non-Discounted", "hslr"): ("784212", "FDEBD0", "E67E22"),
            ("Non-Discounted", "lshr"): ("145A32", "D5F5E3", "1E8449"),
        }
        SL = {"hslr": "High Spend · Low Revenue", "lshr": "Low Spend · High Revenue"}
        FMTS = ["@", "@", "#,##0.00", "#,##0.00", '0.00"x"']
        for col, w in zip("ABCDE", [20, 48, 16, 16, 10]):
            pi.column_dimensions[col].width = w
        cr = 1
        if month_label:
            pi.merge_cells(f"A{cr}:E{cr}")
            c = pi.cell(row=cr, column=1, value=f"  {month_label.upper()}")
            c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
            c.fill = PatternFill("solid", start_color="0F172A")
            c.alignment = Alignment(vertical="center")
            pi.row_dimensions[cr].height = 30
            cr += 1

        for cat in ["Discounted", "Non-Discounted"]:
            ins = data_dict.get(cat, {})
            for itype in ["hslr", "lshr"]:
                hdr_bg, row_bg, sec_bg = SC[(cat, itype)]
                pi.merge_cells(f"A{cr}:E{cr}")
                sc = pi.cell(row=cr, column=1, value=f"  {cat}  ·  {SL[itype]}")
                sc.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
                sc.fill = PatternFill("solid", start_color=sec_bg)
                sc.alignment = Alignment(vertical="center")
                pi.row_dimensions[cr].height = 22
                cr += 1
                df_sec = ins.get(itype, pd.DataFrame())
                if df_sec.empty:
                    pi.merge_cells(f"A{cr}:E{cr}")
                    nc2 = pi.cell(row=cr, column=1, value="  No products match this criteria.")
                    nc2.font = Font(italic=True, color="888888", name="Calibri", size=10)
                    nc2.alignment = Alignment(vertical="center")
                    pi.row_dimensions[cr].height = 18
                    cr += 1
                else:
                    for ci, ltext in enumerate(["Product ID", "Product Title", "Spend (INR)", "Revenue (INR)", "ROI"], 1):
                        _hdr(pi.cell(row=cr, column=ci), ltext, hdr_bg, sz=9)
                    pi.row_dimensions[cr].height = 18
                    cr += 1
                    for ri, row in enumerate(df_sec.itertuples(index=False)):
                        bg2 = row_bg if ri % 2 == 0 else "FDFEFE"
                        for ci, (cv, fmt2) in enumerate(zip(row, FMTS), 1):
                            c2 = pi.cell(row=cr, column=ci, value=cv)
                            c2.number_format = fmt2
                            c2.font  = Font(name="Calibri", size=10)
                            c2.fill  = PatternFill("solid", start_color=bg2)
                            c2.alignment = Alignment(vertical="center", wrap_text=(ci == 2))
                            c2.border = Border(
                                left=Side(style="thin", color="DDDDDD"),
                                right=Side(style="thin", color="DDDDDD"),
                                top=Side(style="thin", color="DDDDDD"),
                                bottom=Side(style="thin", color="DDDDDD"),
                            )
                        pi.row_dimensions[cr].height = 16
                        cr += 1
                    t_sp  = df_sec["Spend"].sum()
                    t_rv  = df_sec["Revenue"].sum()
                    t_roi = round(t_rv / t_sp, 4) if t_sp else 0
                    for ci, (cv, fmt2) in enumerate(
                        zip([f"{len(df_sec)} products", "TOTAL", t_sp, t_rv, t_roi], FMTS), 1
                    ):
                        _tot(pi.cell(row=cr, column=ci), cv, fmt2, hdr_bg)
                    pi.row_dimensions[cr].height = 18
                    cr += 1
                cr += 1
        pi.freeze_panes = "A2"

    for month in months_ordered:
        md2 = {cat: insights.get((month, cat), {}) for cat in ["Discounted", "Non-Discounted"]}
        _write_insights_sheet(wb, f"Insights {month[:3]}", md2, month)
    _write_insights_sheet(wb, "Overall Insights", overall_insights)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf